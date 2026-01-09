# src/scheduler/task_scheduler.py
import os
import yaml
import random
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from ..models.machine import Machine
# 新增：导入Excel处理库
import openpyxl
from openpyxl import Workbook

class TaskScheduler:
    def __init__(self,
                 machine_config_name: str = "machine_config.yaml",
                 task_config_name: str = "task_config.yaml",
                 strategy_config_name: str = "strategy_config.yaml",
                 log_filename: Optional[str] = None,
                 # 优化：移除固定Excel文件名参数，改为动态生成
                 excel_dir: str = "./fragmentation_reports"):
        self.machine_config_path = machine_config_name
        self.task_config_path = task_config_name
        self.strategy_config_path = strategy_config_name

        if log_filename is None:
            self.log_filename = f"scheduler_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        else:
            self.log_filename = log_filename
        self._ensure_log_dir_exists()

        # 优化：动态生成带时间戳的Excel文件名
        self.excel_dir = excel_dir
        self._ensure_excel_dir_exists()
        # 生成带时间戳的Excel文件名，格式：fragmentation_20251210_143025.xlsx
        self.excel_filename = os.path.join(
            self.excel_dir,
            f"fragmentation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        self._init_excel_file()  # 仅创建一次，不会重复创建

        self.machine_config = self.load_yaml(self.machine_config_path)
        self.task_config = self.load_yaml(self.task_config_path)
        self.strategy_config = self.load_yaml(self.strategy_config_path)

        self._init_strategy_params()
        self.machines: List[Machine] = self.init_machines()
        self.task_cache: Dict[str, Dict] = {}
        # 新增：记录调度轮次
        self.schedule_round = 0

    def _ensure_log_dir_exists(self):
        log_dir = os.path.dirname(self.log_filename)
        if log_dir and not os.path.exists(log_dir):
            os.makedirs(log_dir, exist_ok=True)

    # 优化：新增Excel目录创建方法
    def _ensure_excel_dir_exists(self):
        if not os.path.exists(self.excel_dir):
            os.makedirs(self.excel_dir, exist_ok=True)

    # 优化：Excel初始化逻辑（仅创建一次，不会覆盖）
    def _init_excel_file(self):
        # 只有文件不存在时才创建，避免覆盖已有文件
        if not os.path.exists(self.excel_filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "碎片率统计"
            # 设置表头
            ws['A1'] = "调度轮次"
            ws['B1'] = "时间戳"
            ws['C1'] = "平均内存碎片率(%)"
            ws['D1'] = "平均CPU碎片率(%)"
            ws['E1'] = "已分配任务数"
            wb.save(self.excel_filename)
            wb.close()
            print(f"Excel统计文件已创建：{self.excel_filename}")
        else:
            print(f"使用已存在的Excel统计文件：{self.excel_filename}")

    # 优化：保持原有写入逻辑，但文件不会被覆盖
    def write_fragmentation_to_excel(self, mem_avg, cpu_avg, task_count):
        self.schedule_round += 1
        # 容错处理：防止文件被外部程序占用
        try:
            wb = openpyxl.load_workbook(self.excel_filename)
            ws = wb.active
            # 找到下一个空行
            next_row = ws.max_row + 1
            # 写入数据
            ws[f'A{next_row}'] = self.schedule_round
            ws[f'B{next_row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws[f'C{next_row}'] = mem_avg
            ws[f'D{next_row}'] = cpu_avg
            ws[f'E{next_row}'] = task_count
            # 保存文件
            wb.save(self.excel_filename)
            wb.close()
        except Exception as e:
            error_msg = f"写入Excel文件失败：{e}"
            print(error_msg)
            self.write_log_to_file(error_msg)

    def write_log_to_file(self, content: str, add_timestamp: bool = True):
        try:
            if add_timestamp:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                log_line = f"[{timestamp}] {content}\n"
            else:
                log_line = f"{content}\n"
            with open(self.log_filename, 'a', encoding='utf-8') as f:
                f.write(log_line)
        except Exception as e:
            print(f"写入日志文件失败：{e}")

    def load_yaml(self, yaml_filename: str) -> Dict[Any, Any]:
        if not os.path.isabs(yaml_filename):
            current_dir = os.path.dirname(os.path.abspath(__file__))
            yaml_filename = os.path.join(current_dir, yaml_filename)
        full_path = yaml_filename

        if not os.path.exists(full_path):
            error_msg = f"未找到YAML文件：{yaml_filename}（路径：{full_path}）"
            print(error_msg)
            self.write_log_to_file(error_msg)
            raise FileNotFoundError(error_msg)
        try:
            with open(full_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except Exception as e:
            error_msg = f"读取YAML文件{yaml_filename}失败：{e}"
            print(error_msg)
            self.write_log_to_file(error_msg)
            raise

    def _init_strategy_params(self):
        target_weights = self.strategy_config.get('target_weights', {})
        self.memory_frag_weight = target_weights.get('memory_fragment_min', 0.4)
        self.cpu_frag_weight = target_weights.get('cpu_fragment_min', 0.3)
        self.user_distribute_weight = target_weights.get('user_task_distribute', 0.2)
        self.machine_recent_weight = target_weights.get('machine_recent_limit', 0.1)

        user_dist_config = self.strategy_config.get('user_distribute_config', {})
        self.max_user_task_per_machine = user_dist_config.get('max_task_per_user_per_machine', 2)

        machine_recent_config = self.strategy_config.get('machine_recent_config', {})
        self.recent_time_window = machine_recent_config.get('recent_time_window_seconds', 3600)
        self.max_recent_tasks_per_machine = machine_recent_config.get('max_recent_tasks_per_machine', 5)

        frag_config = self.strategy_config.get('fragment_calculation', {})
        self.target_memory_mb = frag_config.get('target_memory_mb', 8192)
        self.target_cpu_cores = frag_config.get('target_cpu_cores', 4)

        total_weight = (
            self.memory_frag_weight +
            self.cpu_frag_weight +
            self.user_distribute_weight +
            self.machine_recent_weight
        )
        if total_weight != 0:
            self.memory_frag_weight /= total_weight
            self.cpu_frag_weight /= total_weight
            self.user_distribute_weight /= total_weight
            self.machine_recent_weight /= total_weight

        # === 精简日志：只保留碎片率相关配置 ===
        weight_log = (
            f"碎片优化策略初始化完成（归一化后）：\n"
            f"  内存碎片最小化权重：{self.memory_frag_weight:.2f}\n"
            f"  CPU碎片最小化权重：{self.cpu_frag_weight:.2f}\n"
            f"  碎片计算基准：内存 {self.target_memory_mb}MB | CPU {self.target_cpu_cores}核"
        )
        self.write_log_to_file(weight_log, add_timestamp=False)

    def init_machines(self) -> List[Machine]:
        count = self.machine_config.get('machine_count', 0)
        cpu = self.machine_config.get('cpu_per_machine', 0)
        memory = self.machine_config.get('memory_per_machine', 0)
        disk = self.machine_config.get('disk_per_machine', 0)
        disabled_ratio = self.machine_config.get('disabled_machine_ratio', 0.0)

        disabled_count = int(count * disabled_ratio)
        enabled_count = count - disabled_count

        machines = []
        for _ in range(enabled_count):
            machines.append(Machine(cpu, memory, disk, is_enabled=True))
        for _ in range(disabled_count):
            machines.append(Machine(cpu, memory, disk, is_enabled=False))
        random.shuffle(machines)

        machine_log = f"机器初始化完成：总数量{len(machines)}台 | 启用{enabled_count}台 | 禁用{disabled_count}台"
        print(machine_log)
        self.write_log_to_file(machine_log)
        return machines

    def _calculate_machine_score(self, machine: Machine, task_cpu: int, task_memory: int,
                                task_disk: int, user_id: str, submit_time) -> float:
        # === 内存碎片得分（预分配后）===
        if machine.free_memory >= task_memory and self.target_memory_mb > 0:
            temp_free = machine.free_memory - task_memory
            temp_frag = (temp_free % self.target_memory_mb) / machine.total_memory * 100
            mem_frag_score = 1 - (temp_frag / 100)
        else:
            mem_frag_rate = machine.calculate_memory_fragmentation_rate(self.target_memory_mb)
            mem_frag_score = 1 - (mem_frag_rate / 100)

        # === CPU碎片得分（预分配后）===
        if machine.free_cpu >= task_cpu and self.target_cpu_cores > 0:
            temp_free_cpu = machine.free_cpu - task_cpu
            temp_cpu_frag = (temp_free_cpu % self.target_cpu_cores) / machine.total_cpu * 100
            cpu_frag_score = 1 - (temp_cpu_frag / 100)
        else:
            cpu_frag_rate = machine.calculate_cpu_fragmentation_rate(self.target_cpu_cores)
            cpu_frag_score = 1 - (cpu_frag_rate / 100)

        # 用户分散 & 近期限流（保留逻辑，但不用于日志）
        user_count = machine.get_user_task_count(user_id)
        if user_count >= self.max_user_task_per_machine:
            user_dist_score = 0.0
        else:
            user_dist_score = (self.max_user_task_per_machine - user_count) / self.max_user_task_per_machine

        recent_count = machine.get_recent_task_count(submit_time, self.recent_time_window)
        if recent_count >= self.max_recent_tasks_per_machine:
            machine_recent_score = 0.0
        else:
            machine_recent_score = (self.max_recent_tasks_per_machine - recent_count) / self.max_recent_tasks_per_machine

        total_score = (
            mem_frag_score * self.memory_frag_weight +
            cpu_frag_score * self.cpu_frag_weight +
            user_dist_score * self.user_distribute_weight +
            machine_recent_score * self.machine_recent_weight
        )
        return round(total_score, 4)

    def strategy_based_allocation(self, task_cpu: int, task_memory: int, task_disk: int,
                                 user_id: str, submit_time) -> Optional[Machine]:
        eligible = []
        cpu_threshold = self.machine_config.get('cpu_overload_threshold', 90)
        overcommit = self.strategy_config.get('resource_overcommit_allowed', False)

        for m in self.machines:
            if not m.is_enabled:
                continue
            if m.get_cpu_usage_rate() >= cpu_threshold:
                continue
            if not overcommit:
                if task_cpu > m.total_cpu or task_memory > m.total_memory or task_disk > m.total_disk:
                    continue
            if m.free_cpu < task_cpu or m.free_memory < task_memory or m.free_disk < task_disk:
                continue
            if m.get_user_task_count(user_id) >= self.max_user_task_per_machine:
                continue
            if m.get_recent_task_count(submit_time, self.recent_time_window) >= self.max_recent_tasks_per_machine:
                continue
            eligible.append(m)

        if not eligible:
            return None

        scores = {m: self._calculate_machine_score(m, task_cpu, task_memory, task_disk, user_id, submit_time)
                  for m in eligible}
        best = max(scores, key=scores.get)
        return best

    def parse_time(self, time_str: str):
        try:
            return datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
        except Exception as e:
            error_msg = f"时间解析失败：{e}"
            print(error_msg)
            self.write_log_to_file(error_msg)
            return None

    def submit_task(self, task_item: Dict) -> bool:
        task_id = task_item.get('task_id')
        cpu = task_item.get('cpu_demand', 0)
        mem = task_item.get('memory_demand_mb', 0)
        disk = task_item.get('disk_demand_gb', 0)
        user_id = task_item.get('user_id', "")
        start_time_str = task_item.get('start_time')
        duration = task_item.get('duration_seconds', 0)

        if cpu <= 0 or mem <= 0 or disk <= 0:
            error_msg = f"任务{task_id}（用户{user_id}）资源非法，分配失败"
            print(error_msg)
            self.write_log_to_file(error_msg)
            return False

        submit_time = self.parse_time(start_time_str)
        if not submit_time:
            error_msg = f"任务{task_id}（用户{user_id}）时间格式非法，分配失败"
            print(error_msg)
            self.write_log_to_file(error_msg)
            return False

        target_machine = self.strategy_based_allocation(cpu, mem, disk, user_id, submit_time)
        if not target_machine:
            error_msg = f"任务{task_id}（用户{user_id}）无可用机器（策略过滤后），分配失败"
            print(error_msg)
            self.write_log_to_file(error_msg)
            return False

        if target_machine.allocate_task(task_id, cpu, mem, disk, user_id, submit_time):
            machine_idx = self.machines.index(target_machine)
            self.task_cache[task_id] = {
                'user_id': user_id,
                'submit_time': submit_time,
                'machine_idx': machine_idx,
                'cpu': cpu,
                'memory': mem,
                'disk': disk
            }
            score = self._calculate_machine_score(target_machine, cpu, mem, disk, user_id, submit_time)
            success_msg = (
                f"任务{task_id}（用户{user_id}）分配成功！\n"
                f"  资源：CPU{cpu}核 | 内存{mem/1024:.1f}GB | 外存{disk}GB\n"
                f"  分配机器：索引{machine_idx} | 机器综合得分：{score}"
            )
            # print(success_msg)
            # self.write_log_to_file(success_msg, add_timestamp=False)
            
            # 新增：每提交一个任务后计算并输出平均碎片率
            self._log_and_save_fragmentation_stats(f"任务{task_id}分配后")
            return True
        else:
            error_msg = f"任务{task_id}（用户{user_id}）分配异常，失败"
            print(error_msg)
            self.write_log_to_file(error_msg)
            return False

    # 新增：核心方法 - 计算、输出、保存碎片率
    def _log_and_save_fragmentation_stats(self, stage: str):
        mem_avg, _ = self.calculate_global_memory_fragmentation_rate()
        cpu_avg, _ = self.calculate_global_cpu_fragmentation_rate()
        task_count = len(self.task_cache)
        
        # 输出平均碎片率
        frag_msg = f"\n【{stage}】平均碎片率统计："
        frag_msg += f"\n  平均内存碎片率：{mem_avg}%"
        frag_msg += f"\n  平均CPU碎片率：{cpu_avg}%"
        frag_msg += f"\n  当前已分配任务数：{task_count}"
        print(frag_msg)
        self.write_log_to_file(frag_msg)
        
        # 写入Excel
        self.write_fragmentation_to_excel(mem_avg, cpu_avg, task_count)

    def submit_all_tasks(self):
        tasks = self.task_config.get('tasks', [])
        if not tasks:
            msg = "无任务可提交"
            print(msg)
            self.write_log_to_file(msg)
            return

        tasks.sort(key=lambda x: self.parse_time(x.get('start_time')) or datetime.min)

        title = f"开始批量提交{len(tasks)}个任务（策略含内存+CPU碎片优化）"
        print("="*60)
        print(title)
        print("="*60)
        self.write_log_to_file("="*60, add_timestamp=False)
        self.write_log_to_file(title, add_timestamp=False)
        self.write_log_to_file("="*60, add_timestamp=False)

        for task in tasks:
            self.submit_task(task)
            # print("-"*40)
            # self.write_log_to_file("-"*40, add_timestamp=False)

        # 新增：所有任务提交完成后输出并保存碎片率
        self._log_and_save_fragmentation_stats("所有任务提交完成")
        
        finish = "所有任务提交完成！"
        print("="*60)
        print(finish)
        print("="*60)
        self.write_log_to_file("="*60, add_timestamp=False)
        self.write_log_to_file(finish, add_timestamp=False)
        self.write_log_to_file("="*60, add_timestamp=False)

    def release_single_task(self, task_id: str) -> bool:
        if task_id not in self.task_cache:
            error_msg = f"任务{task_id}未分配，无法释放"
            print(error_msg)
            self.write_log_to_file(error_msg)
            return False

        info = self.task_cache[task_id]
        machine = self.machines[info['machine_idx']]
        success, log_content = machine.release_task(task_id)
        if success:
            del self.task_cache[task_id]
            # self.write_log_to_file(log_content)
            
            # 新增：释放任务后计算并输出平均碎片率
            self._log_and_save_fragmentation_stats(f"任务{task_id}释放后")
            return True
        else:
            # self.write_log_to_file(log_content)
            return False

    def release_random_tasks(self, release_ratio: float = 0.3):
        if not self.task_cache:
            msg = "无已分配任务，无法随机释放"
            print(msg)
            self.write_log_to_file(msg)
            return

        task_ids = list(self.task_cache.keys())
        release_count = max(1, int(len(task_ids) * release_ratio))
        to_release = random.sample(task_ids, release_count)

        title = f"开始随机释放{len(to_release)}个任务（释放比例：{release_ratio*100}%）"
        print("\n" + "="*60)
        print(title)
        print("="*60)
        self.write_log_to_file("")
        self.write_log_to_file("="*60, add_timestamp=False)
        self.write_log_to_file(title, add_timestamp=False)
        self.write_log_to_file("="*60, add_timestamp=False)

        success = 0
        for tid in to_release:
            if self.release_single_task(tid):
                success += 1
            # print("-"*40)
            self.write_log_to_file("-"*40, add_timestamp=False)

        # 新增：随机释放完成后输出并保存碎片率
        self._log_and_save_fragmentation_stats("随机任务释放完成")
        
        finish = f"随机释放完成！成功释放{success}个任务，剩余任务{len(self.task_cache)}个"
        print(finish)
        print("="*60 + "\n")
        self.write_log_to_file(finish, add_timestamp=False)
        self.write_log_to_file("="*60, add_timestamp=False)
        self.write_log_to_file("")

    def show_task_status(self):
        if not self.task_cache:
            msg = "无已分配任务"
            print(msg)
            self.write_log_to_file(msg)
            return

        title = f"当前已分配任务总数：{len(self.task_cache)}"
        print("\n" + "="*80)
        print(title)
        print("="*80)
        self.write_log_to_file("")
        self.write_log_to_file("="*80, add_timestamp=False)
        self.write_log_to_file(title, add_timestamp=False)
        self.write_log_to_file("="*80, add_timestamp=False)

        for tid, info in self.task_cache.items():
            msg = (
                f"任务ID：{tid} | 用户ID：{info['user_id']}\n"
                f"  分配机器：索引{info['machine_idx']} | 资源：CPU{info['cpu']}核 | 内存{info['memory']/1024:.1f}GB | 外存{info['disk']}GB\n"
                f"  提交时间：{info['submit_time'].strftime('%Y-%m-%d %H:%M:%S')}"
            )
            # print(msg)
            # self.write_log_to_file(msg, add_timestamp=False)

        # print("="*80 + "\n")
        self.write_log_to_file("="*80, add_timestamp=False)
        self.write_log_to_file("")

    def show_strategy_params(self):
        title = "当前调度策略配置参数（仅碎片部分）"
        print("\n" + "="*80)
        print(title)
        print("="*80)
        self.write_log_to_file("")
        self.write_log_to_file("="*80, add_timestamp=False)
        self.write_log_to_file(title, add_timestamp=False)
        self.write_log_to_file("="*80, add_timestamp=False)

        msg = (
            f"碎片优化权重（归一化后）：\n"
            f"  内存碎片最小化权重：{self.memory_frag_weight:.2f}\n"
            f"  CPU碎片最小化权重：{self.cpu_frag_weight:.2f}\n"
            f"\n碎片计算基准：\n"
            f"  目标内存规格：{self.target_memory_mb}MB（{self.target_memory_mb/1024:.1f}GB）\n"
            f"  目标CPU规格：{self.target_cpu_cores}核"
        )
        # print(msg)
        # self.write_log_to_file(msg, add_timestamp=False)

        # print("="*80 + "\n")
        # self.write_log_to_file("="*80, add_timestamp=False)
        # self.write_log_to_file("")

    def calculate_global_memory_fragmentation_rate(self):
        enabled = [m for m in self.machines if m.is_enabled]
        if not enabled:
            return 0.0, []
        rates = [m.calculate_memory_fragmentation_rate(self.target_memory_mb) for m in enabled]
        avg = round(sum(rates) / len(rates), 2)
        return avg, rates

    def calculate_global_cpu_fragmentation_rate(self):
        enabled = [m for m in self.machines if m.is_enabled]
        if not enabled:
            return 0.0, []
        rates = [m.calculate_cpu_fragmentation_rate(self.target_cpu_cores) for m in enabled]
        avg = round(sum(rates) / len(rates), 2)
        return avg, rates

    def print_fragmentation_stats(self, percentiles=[50, 90]):
        # === 内存碎片 ===
        mem_avg, mem_rates = self.calculate_global_memory_fragmentation_rate()
        # === CPU碎片 ===
        cpu_avg, cpu_rates = self.calculate_global_cpu_fragmentation_rate()
        enabled_count = len([m for m in self.machines if m.is_enabled])

        title = "内存与CPU碎片率统计（用于策略对比）"
        print("\n" + "="*80)
        print(title)
        print("="*80)
        self.write_log_to_file("")
        self.write_log_to_file("="*80, add_timestamp=False)
        self.write_log_to_file(title, add_timestamp=False)
        self.write_log_to_file("="*80, add_timestamp=False)

        stats = (
            f"统计范围：{enabled_count} 台启用节点\n"
            f"碎片计算基准：内存 {self.target_memory_mb}MB | CPU {self.target_cpu_cores}核\n"
            f"平均内存碎片率：{mem_avg}%\n"
            f"平均CPU碎片率：{cpu_avg}%\n"
            f"各节点内存碎片率列表：{mem_rates}\n"
            f"各节点CPU碎片率列表：{cpu_rates}"
        )
        print(stats)
        average = sum(cpu_rates) / len(cpu_rates)
        print(f"平均碎片率：{average}")
        self.write_log_to_file(stats, add_timestamp=False)

        if mem_rates:
            sorted_mem = sorted(mem_rates)
            for p in percentiles:
                idx = min(int((p / 100) * len(sorted_mem)), len(sorted_mem) - 1)
                val = sorted_mem[idx]
                msg = f"  内存 {p}%分位碎片率：{val}%"
                print(msg)
                self.write_log_to_file(msg, add_timestamp=False)

        if cpu_rates:
            sorted_cpu = sorted(cpu_rates)
            for p in percentiles:
                idx = min(int((p / 100) * len(sorted_cpu)), len(sorted_cpu) - 1)
                val = sorted_cpu[idx]
                msg = f"  CPU {p}%分位碎片率：{val}%"
                print(msg)
                self.write_log_to_file(msg, add_timestamp=False)

        # print("="*80 + "\n")
        # self.write_log_to_file("="*80, add_timestamp=False)
        # self.write_log_to_file("")

    def show_all_machines_status(self):
        title = f"所有机器状态（总数：{len(self.machines)}台）"
        print("\n" + "="*80)
        print(title)
        print("="*80)
        self.write_log_to_file("")
        self.write_log_to_file("="*80, add_timestamp=False)
        self.write_log_to_file(title, add_timestamp=False)
        self.write_log_to_file("="*80, add_timestamp=False)

        enabled = sum(1 for m in self.machines if m.is_enabled)
        disabled = len(self.machines) - enabled
        stats = (
            f"节点统计：启用{enabled}台 | 禁用{disabled}台\n"
            f"碎片计算基准：内存 {self.target_memory_mb}MB | CPU {self.target_cpu_cores}核"
        )
        print(stats)
        self.write_log_to_file(stats, add_timestamp=False)

        print("-"*80)
        self.write_log_to_file("-"*80, add_timestamp=False)

        for i, m in enumerate(self.machines):
            msg = f"机器{i}：{m.__str__()}"
            print(msg)
            self.write_log_to_file(msg, add_timestamp=False)

        print("="*80 + "\n")
        self.write_log_to_file("="*80, add_timestamp=False)
        self.write_log_to_file("")

    # 以下两个验证方法保留逻辑，但因不涉及碎片，此处略去修改（实际可保留原样）
    def verify_user_distribute_effect(self):
        pass  # 实际保留原逻辑，此处省略以节省篇幅

    def verify_machine_recent_limit_effect(self):
        pass  # 实际保留原逻辑，此处省略以节省篇幅